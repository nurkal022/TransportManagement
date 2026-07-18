from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
from sqlalchemy import or_, and_, text, inspect, func
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import requests

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///transport.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)

class CargoType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    routes = db.relationship('Route', backref='cargo_type', lazy=True)

class Vehicle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(20), nullable=False)
    model = db.Column(db.String(100), nullable=False)
    year = db.Column(db.Integer)
    status = db.Column(db.String(20), default='active')  # active, repair, inactive
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    routes = db.relationship('Route', backref='vehicle', lazy=True)
    driver = db.relationship('Driver', backref='vehicle', uselist=False)

class Driver(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact = db.Column(db.String(20))
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicle.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Отношения
    routes = db.relationship('Route', backref='driver')

    @property
    def current_route(self):
        # Получаем самый последний маршрут водителя
        return Route.query.filter_by(driver_id=self.id).order_by(Route.date.desc()).first()

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'contact': self.contact,
            'vehicle': self.vehicle.number if self.vehicle else None
        }

class Route(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    start_point = db.Column(db.String(200), nullable=False)
    end_point = db.Column(db.String(200), nullable=False)
    date = db.Column(db.DateTime, nullable=False)
    estimated_arrival = db.Column(db.DateTime)
    actual_arrival = db.Column(db.DateTime)
    cargo_type_id = db.Column(db.Integer, db.ForeignKey('cargo_type.id'))
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicle.id'))
    driver_id = db.Column(db.Integer, db.ForeignKey('driver.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    cargo_weight = db.Column(db.Float)
    weight_unit = db.Column(db.String(10), default='t', nullable=False)  # фиксируем тонны
    fuel = db.Column(db.Float)  # литры
    cargo_volume = db.Column(db.Float)
    trips_count = db.Column(db.Integer, default=1)  # Количество рейсов
    invoice_paid = db.Column(db.Boolean, default=False, nullable=False)  # Статус счета
    notes = db.Column(db.Text)      # Дополнительные заметки
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'start_point': self.start_point,
            'end_point': self.end_point,
            'date': self.date.date().isoformat() if self.date else None,
            'estimated_arrival': self.estimated_arrival.isoformat() if self.estimated_arrival else None,
            'actual_arrival': self.actual_arrival.isoformat() if self.actual_arrival else None,
            'driver': self.driver.name if self.driver else None,
            'vehicle': f"{self.vehicle.number} - {self.vehicle.model}" if self.vehicle else None,
            'driver_id': self.driver_id,
            'vehicle_id': self.vehicle_id,
            'cargo_type_id': self.cargo_type_id,
            'cargo_type': self.cargo_type.name if self.cargo_type else None,
            'cargo_weight': self.cargo_weight,
            'weight_unit': self.weight_unit,
            'fuel': self.fuel,
            'cargo_volume': self.cargo_volume,
            'trips_count': self.trips_count,
            'invoice_paid': self.invoice_paid,
            'notes': self.notes
        }

class VehicleActivity(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    vehicle_id = db.Column(db.Integer, db.ForeignKey('vehicle.id'), nullable=False)
    activity_type = db.Column(db.String(100), nullable=False)  # Тип активности (ремонт, обслуживание и т.д.)
    description = db.Column(db.Text, nullable=False)  # Описание работы
    cost = db.Column(db.Float)  # Стоимость
    performed_at = db.Column(db.DateTime, nullable=False)  # Дата выполнения
    performed_by = db.Column(db.String(100))  # Кто выполнял
    notes = db.Column(db.Text)  # Дополнительные заметки
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    vehicle = db.relationship('Vehicle', backref='activities')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- Переключение аккаунтов (доступно любому вошедшему пользователю) ---
@app.context_processor
def inject_switch_accounts():
    """Прокидывает в шаблоны список аккаунтов для быстрого переключения."""
    if current_user.is_authenticated:
        return {'switch_accounts': User.query.order_by(User.username).all()}
    return {'switch_accounts': []}

# Routes
@app.route('/')
@login_required
def index():
    # Основная статистика
    total_routes = Route.query.filter_by(user_id=current_user.id).count()
    total_drivers = Driver.query.filter_by(user_id=current_user.id).count()
    total_vehicles = Vehicle.query.filter_by(user_id=current_user.id).count()
    total_cargo_types = CargoType.query.count()

    # Статистика транспорта
    active_vehicles = Vehicle.query.filter_by(user_id=current_user.id, status='active').count()
    repair_vehicles = Vehicle.query.filter_by(user_id=current_user.id, status='repair').count()
    inactive_vehicles = Vehicle.query.filter_by(user_id=current_user.id, status='inactive').count()

    # Статистика маршрутов
    paid_invoices = Route.query.filter_by(user_id=current_user.id, invoice_paid=True).count()
    unpaid_invoices = total_routes - paid_invoices

    # Статистика за последние 30 дней
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    recent_routes = Route.query.filter(Route.user_id == current_user.id, Route.created_at >= thirty_days_ago).count()

    # Свободные водители
    available_drivers_count = Driver.query.filter_by(user_id=current_user.id, vehicle_id=None).count()

    # Получаем последние маршруты для таблицы
    recent_routes_list = Route.query.filter_by(user_id=current_user.id).order_by(Route.created_at.desc()).limit(5).all()
    
    return render_template('index.html',
                         total_routes=total_routes,
                         total_drivers=total_drivers,
                         total_vehicles=total_vehicles,
                         total_cargo_types=total_cargo_types,
                         active_vehicles=active_vehicles,
                         repair_vehicles=repair_vehicles,
                         inactive_vehicles=inactive_vehicles,
                         paid_invoices=paid_invoices,
                         unpaid_invoices=unpaid_invoices,
                         recent_routes_count=recent_routes,
                         available_drivers_count=available_drivers_count,
                         recent_routes_list=recent_routes_list)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('index'))
        
        flash('Неверное имя пользователя или пароль', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/switch-account/<int:user_id>')
@login_required
def switch_account(user_id):
    """Быстрое переключение между аккаунтами без пароля (для любого вошедшего)."""
    target = User.query.get_or_404(user_id)
    login_user(target)
    flash(f'Вы вошли как «{target.username}»', 'success')
    return redirect(request.referrer or url_for('index'))

# --- Сводка «Все данные» (по всем пользователям, только чтение, admin) ---
def _summary_data(date_from, date_to):
    """Собирает агрегированную сводку по маршрутам всех пользователей.

    date_from/date_to — строки 'YYYY-MM-DD' или None. Возвращает dict с
    итогами и разбивками. Данные только читаются.
    """
    conditions = []
    if date_from:
        try:
            conditions.append(Route.date >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            date_from = None
    if date_to:
        try:
            conditions.append(Route.date < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            date_to = None

    base = Route.query.filter(*conditions)
    total_routes = base.count()

    trips, weight, fuel, volume = db.session.query(
        func.coalesce(func.sum(Route.trips_count), 0),
        func.coalesce(func.sum(Route.cargo_weight), 0.0),
        func.coalesce(func.sum(Route.fuel), 0.0),
        func.coalesce(func.sum(Route.cargo_volume), 0.0),
    ).filter(*conditions).one()

    paid = base.filter(Route.invoice_paid == True).count()
    unpaid = total_routes - paid

    # Разбивка по аккаунтам (включая пустые)
    per_account = []
    for u in User.query.order_by(User.username).all():
        uc = conditions + [Route.user_id == u.id]
        cnt = Route.query.filter(*uc).count()
        t, w, f = db.session.query(
            func.coalesce(func.sum(Route.trips_count), 0),
            func.coalesce(func.sum(Route.cargo_weight), 0.0),
            func.coalesce(func.sum(Route.fuel), 0.0),
        ).filter(*uc).one()
        per_account.append({'username': u.username, 'routes': cnt,
                            'trips': t, 'weight': w, 'fuel': f})
    per_account.sort(key=lambda a: a['routes'], reverse=True)

    # По типам груза
    cargo_rows = db.session.query(
        func.coalesce(CargoType.name, '—'),
        func.count(Route.id),
        func.coalesce(func.sum(Route.cargo_weight), 0.0),
    ).select_from(Route).outerjoin(CargoType, Route.cargo_type_id == CargoType.id) \
     .filter(*conditions).group_by(CargoType.name) \
     .order_by(func.count(Route.id).desc()).all()
    by_cargo = [{'name': r[0], 'routes': r[1], 'weight': r[2]} for r in cargo_rows]

    # По машинам (топ-15)
    veh_rows = db.session.query(
        Vehicle.number, Vehicle.model,
        func.count(Route.id),
        func.coalesce(func.sum(Route.cargo_weight), 0.0),
    ).select_from(Route).join(Vehicle, Route.vehicle_id == Vehicle.id) \
     .filter(*conditions).group_by(Vehicle.id) \
     .order_by(func.count(Route.id).desc()).limit(15).all()
    by_vehicle = [{'number': r[0], 'model': r[1], 'routes': r[2], 'weight': r[3]} for r in veh_rows]

    # Помесячно
    month_rows = db.session.query(
        func.strftime('%Y-%m', Route.date),
        func.count(Route.id),
        func.coalesce(func.sum(Route.cargo_weight), 0.0),
    ).filter(*conditions).group_by(func.strftime('%Y-%m', Route.date)) \
     .order_by(func.strftime('%Y-%m', Route.date).desc()).all()
    by_month = [{'month': r[0], 'routes': r[1], 'weight': r[2]} for r in month_rows]

    return {
        'date_from': date_from or '',
        'date_to': date_to or '',
        'totals': {
            'routes': total_routes, 'trips': trips, 'weight': weight,
            'fuel': fuel, 'volume': volume, 'paid': paid, 'unpaid': unpaid,
        },
        'per_account': per_account,
        'by_cargo': by_cargo,
        'by_vehicle': by_vehicle,
        'by_month': by_month,
    }

@app.route('/summary')
@login_required
def summary():
    data = _summary_data(request.args.get('date_from'), request.args.get('date_to'))
    return render_template('summary.html', **data)

@app.route('/summary/export.xlsx')
@login_required
def summary_export():
    data = _summary_data(request.args.get('date_from'), request.args.get('date_to'))

    output = BytesIO()
    wb = Workbook()
    header_font = Font(bold=True)

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws):
        for column_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    t = data['totals']
    ws = wb.active
    ws.title = "Итоги"
    ws.append(["Показатель", "Значение"])
    style_header(ws)
    period = f"{data['date_from'] or 'начало'} — {data['date_to'] or 'сейчас'}"
    for row in [
        ("Период", period),
        ("Маршрутов", t['routes']),
        ("Рейсов", t['trips']),
        ("Вес груза (т)", round(t['weight'], 2)),
        ("Топливо (л)", round(t['fuel'], 2)),
        ("Объём", round(t['volume'], 2)),
        ("Счета оплачены", t['paid']),
        ("Счета не оплачены", t['unpaid']),
    ]:
        ws.append(list(row))
    autosize(ws)

    ws = wb.create_sheet("По аккаунтам")
    ws.append(["Аккаунт", "Маршрутов", "Рейсов", "Вес (т)", "Топливо (л)"])
    style_header(ws)
    for a in data['per_account']:
        ws.append([a['username'], a['routes'], a['trips'], round(a['weight'], 2), round(a['fuel'], 2)])
    autosize(ws)

    ws = wb.create_sheet("По типам груза")
    ws.append(["Тип груза", "Маршрутов", "Вес (т)"])
    style_header(ws)
    for c in data['by_cargo']:
        ws.append([c['name'], c['routes'], round(c['weight'], 2)])
    autosize(ws)

    ws = wb.create_sheet("По машинам")
    ws.append(["Номер", "Модель", "Маршрутов", "Вес (т)"])
    style_header(ws)
    for v in data['by_vehicle']:
        ws.append([v['number'], v['model'], v['routes'], round(v['weight'], 2)])
    autosize(ws)

    ws = wb.create_sheet("Помесячно")
    ws.append(["Месяц", "Маршрутов", "Вес (т)"])
    style_header(ws)
    for m in data['by_month']:
        ws.append([m['month'], m['routes'], round(m['weight'], 2)])
    autosize(ws)

    wb.save(output)
    output.seek(0)
    filename = f"summary_{data['date_from'] or 'all'}_{data['date_to'] or 'all'}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Cargo Types
@app.route('/cargo_types')
@login_required
def cargo_types():
    # Типы груза общие для всех аккаунтов
    cargo_types = CargoType.query.order_by(CargoType.name).all()
    return render_template('cargo_types.html', cargo_types=cargo_types)

@app.route('/cargo_types/add', methods=['POST'])
@login_required
def cargo_types_add():
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        
        cargo_type = CargoType(name=name, description=description, user_id=current_user.id)
        db.session.add(cargo_type)
        db.session.commit()
        
        flash('Тип груза успешно добавлен', 'success')
        return redirect(url_for('cargo_types'))

@app.route('/api/cargo_types/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def cargo_type_api(id):
    # Типы груза общие — доступны любому вошедшему
    cargo_type = CargoType.query.get_or_404(id)
    
    if request.method == 'GET':
        return jsonify({
            'id': cargo_type.id,
            'name': cargo_type.name,
            'description': cargo_type.description
        })
    
    elif request.method == 'PUT':
        data = request.get_json()
        cargo_type.name = data['name']
        cargo_type.description = data.get('description')
        db.session.commit()
        return jsonify({'message': 'Updated successfully'})
    
    elif request.method == 'DELETE':
        db.session.delete(cargo_type)
        db.session.commit()
        return jsonify({'message': 'Deleted successfully'})

# Vehicles
@app.route('/vehicles')
@login_required
def vehicles():
    vehicles = Vehicle.query.filter_by(user_id=current_user.id).all()
    return render_template('vehicles.html', vehicles=vehicles)

@app.route('/vehicles/add', methods=['POST'])
@login_required
def vehicles_add():
    if request.method == 'POST':
        number = request.form.get('number')
        model = request.form.get('model')
        year = request.form.get('year')
        status = request.form.get('status', 'active')
        
        vehicle = Vehicle(number=number, model=model, year=year, status=status, user_id=current_user.id)
        db.session.add(vehicle)
        db.session.commit()
        
        flash('Транспорт успешно добавлен', 'success')
        return redirect(url_for('vehicles'))

@app.route('/api/vehicles/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def vehicle_api(id):
    vehicle = Vehicle.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    
    if request.method == 'GET':
        return jsonify({
            'id': vehicle.id,
            'number': vehicle.number,
            'model': vehicle.model,
            'year': vehicle.year,
            'status': vehicle.status
        })
    
    elif request.method == 'PUT':
        data = request.get_json()
        vehicle.number = data['number']
        vehicle.model = data['model']
        vehicle.year = data['year']
        vehicle.status = data.get('status', vehicle.status)
        db.session.commit()
        return jsonify({'message': 'Updated successfully'})
    
    elif request.method == 'DELETE':
        db.session.delete(vehicle)
        db.session.commit()
        return jsonify({'message': 'Deleted successfully'})

# Drivers
@app.route('/drivers')
@login_required
def drivers():
    drivers = Driver.query.filter_by(user_id=current_user.id).all()
    available_vehicles = Vehicle.query.filter_by(user_id=current_user.id, status='active').all()
    
    return render_template('drivers.html',
                         drivers=drivers,
                         available_vehicles=available_vehicles)

@app.route('/drivers/add', methods=['POST'])
@login_required
def drivers_add():
    if request.method == 'POST':
        name = request.form.get('name')
        contact = request.form.get('contact')
        vehicle_id = request.form.get('vehicle_id')
        
        driver = Driver(name=name, contact=contact, vehicle_id=vehicle_id, user_id=current_user.id)
        db.session.add(driver)
        db.session.commit()
        
        flash('Водитель успешно добавлен', 'success')
        return redirect(url_for('drivers'))

@app.route('/api/drivers/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def driver_api(id):
    driver = Driver.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    
    if request.method == 'GET':
        return jsonify({
            'id': driver.id,
            'name': driver.name,
            'contact': driver.contact,
            'vehicle_id': driver.vehicle_id
        })
    
    elif request.method == 'PUT':
        data = request.get_json()
        driver.name = data['name']
        driver.contact = data['contact']
        driver.vehicle_id = data.get('vehicle_id')
        db.session.commit()
        return jsonify({'message': 'Updated successfully'})
    
    elif request.method == 'DELETE':
        db.session.delete(driver)
        db.session.commit()
        return jsonify({'message': 'Deleted successfully'})

@app.route('/api/drivers/<int:id>/history')
@login_required
def driver_history(id):
    driver = Driver.query.get_or_404(id)
    routes = Route.query.filter_by(driver_id=id).order_by(Route.date.desc()).all()
    return jsonify([{
        'name': route.name,
        'date': route.date.isoformat(),
        'vehicle_number': route.vehicle.number if route.vehicle else None,
        'vehicle_model': route.vehicle.model if route.vehicle else None,
        'status': route.status
    } for route in routes])

# Routes
@app.route('/routes')
@login_required
def routes():
    # Получаем параметры фильтрации
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    driver_id = request.args.get('driver_id')
    vehicle_id = request.args.get('vehicle_id')
    cargo_type_id = request.args.get('cargo_type_id')
    start_point = request.args.get('start_point')
    end_point = request.args.get('end_point')
    
    # Базовый запрос с дефолтным окном 90 дней, если даты не заданы
    query = Route.query
    default_window_start = datetime.utcnow() - timedelta(days=90)
    
    # Применяем фильтры по точкам отправления и назначения
    if start_point:
        query = query.filter(Route.start_point.ilike(f'%{start_point}%'))
    
    if end_point:
        query = query.filter(Route.end_point.ilike(f'%{end_point}%'))
    
    # Применяем фильтры дат
    if start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Route.date.between(start, end))
    elif not start_date and not end_date:
        # дефолтно последние 90 дней
        query = query.filter(Route.date >= default_window_start)
    
    if driver_id:
        query = query.filter(Route.driver_id == driver_id)
    
    if vehicle_id:
        query = query.filter(Route.vehicle_id == vehicle_id)
    
    if cargo_type_id:
        query = query.filter(Route.cargo_type_id == cargo_type_id)
    
    routes = query.order_by(Route.date.desc()).limit(50).all()
    
    drivers = Driver.query.all()
    vehicles = Vehicle.query.filter_by(status='active').all()
    cargo_types = CargoType.query.all()
    
    total_routes = len(routes)
    total_trips = sum(route.trips_count for route in routes)
    
    current_filters = {
        'start_date': start_date,
        'end_date': end_date,
        'driver_id': driver_id,
        'vehicle_id': vehicle_id,
        'cargo_type_id': cargo_type_id,
        'start_point': start_point,
        'end_point': end_point
    }
    
    routes_data = [route.to_dict() for route in routes]
    
    return render_template('routes.html',
                         routes=routes,
                         drivers=drivers,
                         vehicles=vehicles,
                         cargo_types=cargo_types,
                         current_filters=current_filters,
                         total_routes=total_routes,
                         total_trips=total_trips,
                         routes_data=routes_data,
                         default_window_start=default_window_start.date().isoformat())

@app.route('/routes/add', methods=['POST'])
@login_required
def routes_add():
    if request.method == 'POST':
        name = request.form.get('name')
        start_point = request.form.get('start_point')
        end_point = request.form.get('end_point')
        date = datetime.strptime(request.form.get('date'), '%Y-%m-%d')
        driver_id = request.form.get('driver_id')
        vehicle_id = request.form.get('vehicle_id')
        cargo_type_id = request.form.get('cargo_type_id')
        cargo_weight_raw = request.form.get('cargo_weight')
        cargo_weight = float(cargo_weight_raw) if cargo_weight_raw else None
        weight_unit = request.form.get('weight_unit', 't')
        cargo_volume_raw = request.form.get('cargo_volume')
        cargo_volume = float(cargo_volume_raw) if cargo_volume_raw else None
        trips_count = int(request.form.get('trips_count', 1))
        notes = request.form.get('notes')
        fuel = request.form.get('fuel')
        invoice_paid = request.form.get('invoice_paid') == 'on'
        
        # Расчет ожидаемого времени прибытия (примерно)
        estimated_arrival = date + timedelta(hours=24)  # По умолчанию 24 часа
        
        route = Route(
            name=name,
            start_point=start_point,
            end_point=end_point,
            date=date,
            estimated_arrival=estimated_arrival,
            driver_id=driver_id,
            vehicle_id=vehicle_id,
            cargo_type_id=cargo_type_id,
            user_id=current_user.id,
            cargo_weight=cargo_weight,
            weight_unit=weight_unit or 't',
            fuel=float(fuel) if fuel else None,
            cargo_volume=cargo_volume,
            trips_count=trips_count,
            invoice_paid=invoice_paid,
            notes=notes
        )
        
        db.session.add(route)
        db.session.commit()
        
        flash('Маршрут успешно добавлен', 'success')
        return redirect(url_for('routes'))

@app.route('/api/routes/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def route_api(id):
    route = Route.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    
    if request.method == 'GET':
        return jsonify({
            'id': route.id,
            'name': route.name,
            'start_point': route.start_point,
            'end_point': route.end_point,
            'date': route.date.strftime('%Y-%m-%d'),
            'driver_id': route.driver_id,
            'vehicle_id': route.vehicle_id,
            'cargo_type_id': route.cargo_type_id,
            'cargo_weight': route.cargo_weight,
            'weight_unit': route.weight_unit,
            'fuel': route.fuel,
            'cargo_volume': route.cargo_volume,
            'trips_count': route.trips_count,
            'invoice_paid': route.invoice_paid,
            'notes': route.notes
        })
    
    elif request.method == 'PUT':
        data = request.get_json()
        route.name = data['name']
        route.start_point = data['start_point']
        route.end_point = data['end_point']
        route.date = datetime.strptime(data['date'], '%Y-%m-%d')
        route.driver_id = data.get('driver_id')
        route.vehicle_id = data.get('vehicle_id')
        route.cargo_type_id = data['cargo_type_id']
        route.cargo_weight = float(data.get('cargo_weight', 0))
        route.weight_unit = data.get('weight_unit', 't')
        fuel_value = data.get('fuel')
        route.fuel = float(fuel_value) if fuel_value not in (None, '',) else None
        route.cargo_volume = float(data.get('cargo_volume', 0))
        route.trips_count = int(data.get('trips_count', 1))
        route.invoice_paid = bool(data.get('invoice_paid', False))
        route.notes = data.get('notes', '')
        db.session.commit()
        return jsonify({'success': True, 'message': 'Updated successfully'})
    
    elif request.method == 'DELETE':
        db.session.delete(route)
        db.session.commit()
        return jsonify({'message': 'Deleted successfully'})

@app.route('/api/routes/<int:id>/details')
@login_required
def route_details(id):
    route = Route.query.get_or_404(id)
    return jsonify({
        'id': route.id,
        'name': route.name,
        'start_point': route.start_point,
        'end_point': route.end_point,
        'date': route.date.isoformat() if route.date else None,
        'driver': route.driver.name if route.driver else None,
        'vehicle': route.vehicle.number if route.vehicle else None,
        'cargo_type': route.cargo_type.name if route.cargo_type else None,
        'cargo_weight': route.cargo_weight,
        'weight_unit': route.weight_unit,
        'fuel': route.fuel,
        'cargo_volume': route.cargo_volume,
        'trips_count': route.trips_count,
        'invoice_paid': route.invoice_paid,
        'notes': route.notes,
        'estimated_arrival': route.estimated_arrival.isoformat() if route.estimated_arrival else None,
        'actual_arrival': route.actual_arrival.isoformat() if route.actual_arrival else None
    })

@app.route('/api/routes/<int:id>', methods=['DELETE'])
@login_required
def delete_route_api(id):
    route = Route.query.get_or_404(id)
    db.session.delete(route)
    db.session.commit()
    return jsonify({'message': 'Маршрут успешно удален'})

@app.route('/api/routes/<int:id>/invoice', methods=['PUT'])
@login_required
def route_invoice_api(id):
    route = Route.query.get_or_404(id)
    data = request.get_json() or {}
    route.invoice_paid = bool(data.get('invoice_paid', False))
    db.session.commit()
    return jsonify({'success': True, 'invoice_paid': route.invoice_paid})

# Vehicle Activities
@app.route('/vehicles/<int:vehicle_id>/activities')
@login_required
def vehicle_activities(vehicle_id):
    vehicle = Vehicle.query.filter_by(id=vehicle_id, user_id=current_user.id).first_or_404()
    activities = VehicleActivity.query.filter_by(vehicle_id=vehicle_id).order_by(VehicleActivity.performed_at.desc()).all()
    return render_template('vehicle_activities.html', vehicle=vehicle, activities=activities)

@app.route('/api/vehicles/<int:vehicle_id>/activities', methods=['GET', 'POST'])
@login_required
def vehicle_activities_api(vehicle_id):
    vehicle = Vehicle.query.filter_by(id=vehicle_id, user_id=current_user.id).first_or_404()

    if request.method == 'GET':
        activities = VehicleActivity.query.filter_by(vehicle_id=vehicle_id).order_by(VehicleActivity.performed_at.desc()).all()
        return jsonify([{
            'id': a.id,
            'activity_type': a.activity_type,
            'description': a.description,
            'cost': a.cost,
            'performed_at': a.performed_at.isoformat() if a.performed_at else None,
            'performed_by': a.performed_by,
            'notes': a.notes
        } for a in activities])

    elif request.method == 'POST':
        data = request.get_json()
        activity = VehicleActivity(
            vehicle_id=vehicle_id,
            activity_type=data['activity_type'],
            description=data['description'],
            cost=float(data.get('cost', 0)) if data.get('cost') else None,
            performed_at=datetime.strptime(data['performed_at'], '%Y-%m-%dT%H:%M'),
            performed_by=data.get('performed_by'),
            notes=data.get('notes')
        )
        db.session.add(activity)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Активность добавлена'})

@app.route('/api/vehicles/activities/<int:activity_id>', methods=['DELETE'])
@login_required
def delete_vehicle_activity(activity_id):
    activity = VehicleActivity.query.get_or_404(activity_id)
    db.session.delete(activity)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Активность удалена'})

@app.route('/api/routes/stats')
@login_required
def routes_stats():
    # Статистика по маршрутам
    total_routes = Route.query.count()
    
    # Статистика по водителям
    top_drivers = db.session.query(
        Driver.name,
        db.func.count(Route.id).label('total_routes')
    ).join(Route).group_by(Driver.id).order_by(db.desc('total_routes')).limit(5).all()
    
    # Статистика по транспорту
    vehicle_stats = db.session.query(
        Vehicle.number,
        db.func.count(Route.id).label('total_routes')
    ).join(Route).group_by(Vehicle.id).all()
    
    return jsonify({
        'total_routes': total_routes,
        'top_drivers': [{'name': d[0], 'routes': d[1]} for d in top_drivers],
        'vehicle_stats': [{'number': v[0], 'routes': v[1]} for v in vehicle_stats]
    })

# API Routes
@app.route('/api/routes')
@login_required
def routes_api():
    load_all = request.args.get('load_all') == '1'
    before_date = request.args.get('before_date')
    page_limit = int(request.args.get('limit', 50))

    query = build_routes_query_from_request(include_default_window=not load_all)

    if load_all:
        routes = query.order_by(Route.date.desc()).all()
        has_more = False
        next_before_date = None
    else:
        if before_date:
            try:
                bd = datetime.strptime(before_date, '%Y-%m-%d')
                query = query.filter(Route.date < bd)
            except ValueError:
                pass
        routes = query.order_by(Route.date.desc()).limit(page_limit + 1).all()
        has_more = len(routes) > page_limit
        routes = routes[:page_limit]
        next_before_date = routes[-1].date.date().isoformat() if has_more and routes[-1].date else None

    return jsonify({
        'routes': [{
            'id': route.id,
            'name': route.name,
            'start_point': route.start_point,
            'end_point': route.end_point,
            'date': route.date.date().isoformat() if route.date else None,
            'driver': route.driver.name if route.driver else None,
            'vehicle': f"{route.vehicle.number} - {route.vehicle.model}" if route.vehicle else None,
            'cargo_type': route.cargo_type.name if route.cargo_type else None,
            'cargo_weight': route.cargo_weight,
            'weight_unit': route.weight_unit,
            'fuel': route.fuel,
            'trips_count': route.trips_count,
            'invoice_paid': route.invoice_paid,
            'notes': route.notes
        } for route in routes],
        'has_more': has_more,
        'next_before_date': next_before_date
    })

def build_routes_query_from_request(include_default_window=True):
    search = request.args.get('search', '').strip().lower()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    driver_id = request.args.get('driver_id')
    vehicle_id = request.args.get('vehicle_id')
    cargo_type_id = request.args.get('cargo_type_id')
    start_point = request.args.get('start_point', '').strip()
    end_point = request.args.get('end_point', '').strip()
    before_date = request.args.get('before_date')

    query = Route.query.filter_by(user_id=current_user.id)

    if search:
        search_filter = or_(
            Route.name.ilike(f'%{search}%'),
            Route.start_point.ilike(f'%{search}%'),
            Route.end_point.ilike(f'%{search}%'),
            Route.notes.ilike(f'%{search}%'),
            Driver.name.ilike(f'%{search}%'),
            Vehicle.number.ilike(f'%{search}%'),
            Vehicle.model.ilike(f'%{search}%'),
            CargoType.name.ilike(f'%{search}%')
        )
        query = query.join(Driver, Route.driver_id == Driver.id, isouter=True) \
                     .join(Vehicle, Route.vehicle_id == Vehicle.id, isouter=True) \
                     .join(CargoType, Route.cargo_type_id == CargoType.id, isouter=True) \
                    .filter(search_filter)
    
    if start_point:
        query = query.filter(Route.start_point.ilike(f'%{start_point}%'))
    
    if end_point:
        query = query.filter(Route.end_point.ilike(f'%{end_point}%'))
    
    if start_date and end_date:
        try:
            # Попробуем сначала формат DD.MM.YYYY (из формы)
            try:
                start_date_obj = datetime.strptime(start_date, '%d.%m.%Y').date()
                end_date_obj = datetime.strptime(end_date, '%d.%m.%Y').date()
            except ValueError:
                # Если не получилось, попробуем YYYY-MM-DD (из URL параметров)
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Фильтруем по датам без учета времени
            from sqlalchemy import func, and_
            query = query.filter(
                and_(
                    func.date(Route.date) >= start_date_obj,
                    func.date(Route.date) <= end_date_obj
                )
            )
        except ValueError:
            pass
    elif include_default_window and not before_date:
        query = query.filter(Route.date >= (datetime.utcnow() - timedelta(days=90)))

    if before_date:
        try:
            bd = datetime.strptime(before_date, '%Y-%m-%d')
            query = query.filter(Route.date < bd)
        except ValueError:
            pass
    
    if driver_id:
        query = query.filter(Route.driver_id == driver_id)
    
    if vehicle_id:
        query = query.filter(Route.vehicle_id == vehicle_id)
    
    if cargo_type_id:
        query = query.filter(Route.cargo_type_id == cargo_type_id)
    
    return query

@app.route('/routes/export/excel')
@login_required
def export_routes_excel():
    query = build_routes_query_from_request(include_default_window=False)
    routes = query.order_by(Route.date.desc()).all()
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Маршруты"

    headers = [
        "Название", "Откуда", "Куда", "Дата", "Водитель", "Транспорт",
        "Тип груза", "Вес (т)", "Рейсов", "Статус счета", "Примечания"
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in routes:
        ws.append([
            r.name,
            r.start_point,
            r.end_point,
            r.date.date().isoformat() if r.date else "",
            r.driver.name if r.driver else "",
            f"{r.vehicle.number} - {r.vehicle.model}" if r.vehicle else "",
            r.cargo_type.name if r.cargo_type else "",
            r.cargo_weight if r.cargo_weight is not None else "",
            r.trips_count,
            "Да" if r.invoice_paid else "Нет",
            r.notes or ""
        ])

    # Автоширина
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    wb.save(output)
    output.seek(0)
    filename = f"routes_{datetime.utcnow().date().isoformat()}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def get_pdf_font():
    # Сначала проверяем пользовательский шрифт font.ttf
    custom_font_path = os.path.join(app.root_path, 'static', 'fonts', 'font.ttf')
    if os.path.exists(custom_font_path):
        try:
            pdfmetrics.registerFont(TTFont('CustomFont', custom_font_path))
            return 'CustomFont'
        except Exception as e:
            print(f"Не удалось загрузить пользовательский шрифт: {e}")

    # Затем проверяем DejaVuSans.ttf
    dejavu_font_path = os.path.join(app.root_path, 'static', 'fonts', 'DejaVuSans.ttf')
    if os.path.exists(dejavu_font_path):
        try:
            pdfmetrics.registerFont(TTFont('DejaVu', dejavu_font_path))
            return 'DejaVu'
        except Exception as e:
            print(f"Не удалось загрузить DejaVu шрифт: {e}")

    # Если локальных шрифтов нет, пытаемся скачать DejaVu
    try:
        os.makedirs(os.path.dirname(dejavu_font_path), exist_ok=True)
        url = 'https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf'
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            with open(dejavu_font_path, 'wb') as f:
                f.write(response.content)
            pdfmetrics.registerFont(TTFont('DejaVu', dejavu_font_path))
            return 'DejaVu'
    except Exception as e:
        print(f"Не удалось скачать шрифт: {e}")

    # Fallback на Helvetica (не поддерживает кириллицу полностью)
    return 'Helvetica'

@app.route('/routes/export/pdf')
@login_required
def export_routes_pdf():
    # Получаем параметры фильтрации для отчета
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = build_routes_query_from_request(include_default_window=False)
    routes = query.order_by(Route.date.desc()).all()

    output = BytesIO()
    font_name = get_pdf_font()

    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    styles['Normal'].fontName = font_name
    styles['Heading1'].fontName = font_name
    styles['Heading2'].fontName = font_name

    # Создаем стиль для заголовка
    title_style = styles['Heading1']
    title_style.fontSize = 24
    title_style.spaceAfter = 30
    title_style.alignment = 1  # center

    # Создаем стиль для подзаголовков
    subtitle_style = styles['Heading2']
    subtitle_style.fontSize = 14
    subtitle_style.spaceAfter = 20

    # Создаем стиль для обычного текста
    normal_style = styles['Normal']
    normal_style.fontSize = 10

    elements = []

    # Титульная страница
    # Заголовок компании
    company_title_style = styles['Heading1']
    company_title_style.fontSize = 28
    company_title_style.spaceAfter = 20
    company_title_style.alignment = 1  # center
    company_title_style.textColor = colors.darkblue

    elements.append(Paragraph("ТРАНСПОРТНАЯ КОМПАНИЯ", company_title_style))
    elements.append(Paragraph("СИСТЕМА УПРАВЛЕНИЯ ТРАНСПОРТОМ", title_style))
    elements.append(Spacer(1, 40))

    # Основная информация об отчете
    report_info_style = styles['Heading2']
    report_info_style.fontSize = 16
    report_info_style.spaceAfter = 15
    report_info_style.alignment = 1

    elements.append(Paragraph("ОТЧЕТ ПО МАРШРУТАМ", report_info_style))
    elements.append(Spacer(1, 30))

    # Детали отчета в рамке
    from reportlab.platypus import Table, TableStyle

    current_date = datetime.utcnow().strftime('%d.%m.%Y %H:%M')

    # Определяем период отчета
    report_period = "Все данные"
    if start_date and end_date:
        report_period = f"{start_date} - {end_date}"
    elif start_date:
        report_period = f"С {start_date}"
    elif end_date:
        report_period = f"По {end_date}"

    report_details_data = [
        ["Дата генерации:", current_date],
        ["Общее количество маршрутов:", str(len(routes))],
        ["Период отчета:", report_period],
    ]

    report_details_table = Table(report_details_data, colWidths=[120, 200])
    report_details_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    elements.append(report_details_table)
    elements.append(Spacer(1, 30))

    # Статистика в виде красивой таблицы
    paid_count = sum(1 for r in routes if r.invoice_paid)
    unpaid_count = len(routes) - paid_count
    total_weight = sum(r.cargo_weight or 0 for r in routes)
    total_volume = sum(r.cargo_volume or 0 for r in routes)
    total_fuel = sum(r.fuel or 0 for r in routes)
    total_trips = sum(r.trips_count or 1 for r in routes)

    elements.append(Paragraph("ОБЩАЯ СТАТИСТИКА", subtitle_style))

    stats_data = [
        ["Показатель", "Значение"],
        ["Маршруты всего", str(len(routes))],
        ["Оплаченные счета", f"{paid_count} ({paid_count/len(routes)*100:.1f}%)" if routes else "0"],
        ["Неоплаченные счета", f"{unpaid_count} ({unpaid_count/len(routes)*100:.1f}%)" if routes else "0"],
        ["Общий вес груза", f"{total_weight:.1f} т"],
        ["Общий объем груза", f"{total_volume:.1f} м³"],
        ["Общий расход топлива", f"{total_fuel:.1f} л"],
        ["Общее количество рейсов", str(total_trips)],
    ]

    stats_table = Table(stats_data, colWidths=[150, 100])
    stats_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (0, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
    ]))

    elements.append(stats_table)
    elements.append(Spacer(1, 40))

    # Параметры фильтрации (если есть) - показываем в виде таблицы
    search = request.args.get('search', '').strip()
    driver_id = request.args.get('driver_id')
    vehicle_id = request.args.get('vehicle_id')
    cargo_type_id = request.args.get('cargo_type_id')

    filter_info = []
    if search:
        filter_info.append(["Поисковый запрос", search])
    if start_date or end_date:
        date_range = f"{start_date or 'начало'} - {end_date or 'настоящее время'}"
        filter_info.append(["Период", date_range])
    if driver_id:
        driver = Driver.query.get(int(driver_id))
        if driver:
            filter_info.append(["Водитель", driver.name])
    if vehicle_id:
        vehicle = Vehicle.query.get(int(vehicle_id))
        if vehicle:
            filter_info.append(["Транспорт", f"{vehicle.number} - {vehicle.model}"])
    if cargo_type_id:
        cargo_type = CargoType.query.get(int(cargo_type_id))
        if cargo_type:
            filter_info.append(["Тип груза", cargo_type.name])

    if filter_info:
        elements.append(Paragraph("ПРИМЕНЕННЫЕ ФИЛЬТРЫ", subtitle_style))

        filter_table = Table(filter_info, colWidths=[120, 200])
        filter_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(filter_table)
        elements.append(Spacer(1, 20))

    elements.append(PageBreak())

    # Таблица данных
    elements.append(Paragraph("Список маршрутов", subtitle_style))

    # Создаем данные для таблицы
    data = [[
        "Название", "Откуда", "Куда", "Дата", "Водитель", "Транспорт",
        "Тип груза", "Вес (т)", "Рейсов", "Счет", "Примечания"
    ]]

    for r in routes:
        data.append([
            r.name or "",
            r.start_point or "",
            r.end_point or "",
            r.date.date().isoformat() if r.date else "",
            r.driver.name if r.driver else "",
            f"{r.vehicle.number}\n{r.vehicle.model}" if r.vehicle else "",
            r.cargo_type.name if r.cargo_type else "",
            f"{r.cargo_weight:.1f}" if r.cargo_weight is not None else "",
            str(r.trips_count),
            "✓" if r.invoice_paid else "✗",
            r.notes or ""
        ])

    # Создаем таблицу с уменьшенными ширинами колонок для предотвращения наложения
    col_widths = [50, 45, 45, 40, 55, 65, 50, 35, 30, 25, 60]  # Уменьшенные ширины колонок

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 1),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))

    elements.append(table)

    # Нижний колонтитул
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(f"Отчет сгенерирован системой управления транспортом - {current_date}", normal_style))

    doc.build(elements)
    output.seek(0)
    filename = f"routes_report_{datetime.utcnow().date().isoformat()}.pdf"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

@app.route('/statistics')
@login_required
def statistics():
    # Получаем параметры фильтрации
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    driver_id = request.args.get('driver_id')
    vehicle_id = request.args.get('vehicle_id')
    cargo_type_id = request.args.get('cargo_type_id')

    # Начинаем с базового запроса
    query = Route.query.filter_by(user_id=current_user.id)

    # Применяем фильтры по датам
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            from sqlalchemy import func, and_
            query = query.filter(
                and_(
                    func.date(Route.date) >= start_date_obj,
                    func.date(Route.date) <= end_date_obj
                )
            )
        except ValueError:
            pass
    else:
        # По умолчанию за последние 3 месяца
        three_months_ago = datetime.utcnow() - timedelta(days=90)
        query = query.filter(Route.date >= three_months_ago)

    if driver_id:
        query = query.filter(Route.driver_id == driver_id)

    if vehicle_id:
        query = query.filter(Route.vehicle_id == vehicle_id)

    if cargo_type_id:
        query = query.filter(Route.cargo_type_id == cargo_type_id)

    routes = query.all()

    # Рассчитываем метрики
    total_routes = len(routes)
    total_trips = sum(r.trips_count for r in routes)
    total_weight = sum(r.cargo_weight or 0 for r in routes)
    total_volume = sum(r.cargo_volume or 0 for r in routes)
    total_fuel = sum(r.fuel or 0 for r in routes)
    paid_invoices = sum(1 for r in routes if r.invoice_paid)
    unpaid_invoices = total_routes - paid_invoices

    # Статистика по водителям
    driver_stats = {}
    for route in routes:
        if route.driver:
            driver_name = route.driver.name
            if driver_name not in driver_stats:
                driver_stats[driver_name] = {'routes': 0, 'trips': 0, 'weight': 0}
            driver_stats[driver_name]['routes'] += 1
            driver_stats[driver_name]['trips'] += route.trips_count
            driver_stats[driver_name]['weight'] += route.cargo_weight or 0

    # Статистика по транспорту
    vehicle_stats = {}
    for route in routes:
        if route.vehicle:
            vehicle_name = f"{route.vehicle.number} - {route.vehicle.model}"
            if vehicle_name not in vehicle_stats:
                vehicle_stats[vehicle_name] = {'routes': 0, 'trips': 0, 'weight': 0, 'fuel': 0}
            vehicle_stats[vehicle_name]['routes'] += 1
            vehicle_stats[vehicle_name]['trips'] += route.trips_count
            vehicle_stats[vehicle_name]['weight'] += route.cargo_weight or 0
            vehicle_stats[vehicle_name]['fuel'] += route.fuel or 0

    # Статистика по типам груза
    cargo_stats = {}
    for route in routes:
        if route.cargo_type:
            cargo_name = route.cargo_type.name
            if cargo_name not in cargo_stats:
                cargo_stats[cargo_name] = {'routes': 0, 'weight': 0, 'volume': 0}
            cargo_stats[cargo_name]['routes'] += 1
            cargo_stats[cargo_name]['weight'] += route.cargo_weight or 0
            cargo_stats[cargo_name]['volume'] += route.cargo_volume or 0

    # Получаем справочники для фильтров
    drivers = Driver.query.all()
    vehicles = Vehicle.query.all()
    cargo_types = CargoType.query.all()

    current_filters = {
        'start_date': start_date,
        'end_date': end_date,
        'driver_id': driver_id,
        'vehicle_id': vehicle_id,
        'cargo_type_id': cargo_type_id
    }

    return render_template('statistics.html',
                         total_routes=total_routes,
                         total_trips=total_trips,
                         total_weight=total_weight,
                         total_volume=total_volume,
                         total_fuel=total_fuel,
                         paid_invoices=paid_invoices,
                         unpaid_invoices=unpaid_invoices,
                         driver_stats=driver_stats,
                         vehicle_stats=vehicle_stats,
                         cargo_stats=cargo_stats,
                         drivers=drivers,
                         vehicles=vehicles,
                         cargo_types=cargo_types,
                         current_filters=current_filters)

@app.route('/api/monthly-stats')
@login_required
def monthly_stats_api():
    month_str = request.args.get('month')
    if not month_str:
        return jsonify({'error': 'Month parameter is required'}), 400

    try:
        year, month = map(int, month_str.split('-'))
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)

        # Получаем маршруты за выбранный месяц
        routes = Route.query.filter(Route.user_id == current_user.id, Route.date >= start_date, Route.date < end_date).all()

        routes_count = len(routes)
        total_weight = sum(r.cargo_weight or 0 for r in routes)
        total_fuel = sum(r.fuel or 0 for r in routes)
        paid_invoices = sum(1 for r in routes if r.invoice_paid)

        return jsonify({
            'routes_count': routes_count,
            'total_weight': total_weight,
            'total_fuel': total_fuel,
            'paid_invoices': paid_invoices
        })
    except ValueError:
        return jsonify({'error': 'Invalid month format'}), 400

@app.route('/api/stats')
@login_required
def stats_api():
    total_routes = Route.query.count()
    total_drivers = Driver.query.count()
    total_vehicles = Vehicle.query.count()
    
    return jsonify({
        'total_routes': total_routes,
        'total_drivers': total_drivers,
        'total_vehicles': total_vehicles
    })

@app.route('/api/routes/<int:id>', methods=['PUT'])
@login_required
def update_route_api(id):
    route = Route.query.get_or_404(id)
    data = request.json
    
    try:
        # Обновляем основные поля
        route.name = data['name']
        route.start_point = data['start_point']
        route.end_point = data['end_point']
        route.date = datetime.strptime(data['date'], '%Y-%m-%d')
        route.driver_id = data['driver_id']
        route.vehicle_id = data['vehicle_id']
        route.cargo_type_id = data['cargo_type_id']
        route.cargo_weight = float(data['cargo_weight'])
        route.weight_unit = data.get('weight_unit', 't')
        fuel_value = data.get('fuel')
        route.fuel = float(fuel_value) if fuel_value not in (None, '',) else None
        route.cargo_volume = float(data['cargo_volume'])
        route.trips_count = int(data['trips_count'])
        route.invoice_paid = bool(data.get('invoice_paid', False))
        route.notes = data['notes']
        
        # Обновляем расчетное время прибытия
        route.estimated_arrival = route.date + timedelta(hours=24)  # По умолчанию 24 часа
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Маршрут успешно обновлен'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500

# Create all database tables
def ensure_route_columns():
    """Add new columns safely for backward compatibility (SQLite)."""
    if not db.engine.url.drivername.startswith('sqlite'):
        return
    inspector = inspect(db.engine)

    # Обработка таблицы route
    existing_cols = {col['name'] for col in inspector.get_columns('route')}
    with db.engine.begin() as conn:
        if 'fuel' not in existing_cols:
            conn.execute(text("ALTER TABLE route ADD COLUMN fuel REAL"))
        if 'weight_unit' not in existing_cols:
            conn.execute(text("ALTER TABLE route ADD COLUMN weight_unit VARCHAR(10) DEFAULT 't'"))
            conn.execute(text("UPDATE route SET weight_unit='t' WHERE weight_unit IS NULL OR weight_unit='' OR weight_unit='kg'"))
        if 'invoice_paid' not in existing_cols:
            conn.execute(text("ALTER TABLE route ADD COLUMN invoice_paid INTEGER DEFAULT 0"))
            # всем существующим маршрутам ставим галочку
            conn.execute(text("UPDATE route SET invoice_paid=1"))
        if 'user_id' not in existing_cols:
            conn.execute(text("ALTER TABLE route ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1"))
            # Создаем внешний ключ
            try:
                conn.execute(text("ALTER TABLE route ADD CONSTRAINT fk_route_user_id FOREIGN KEY (user_id) REFERENCES user(id)"))
            except Exception as e:
                print(f"Не удалось добавить внешний ключ для route: {e}")

    # Обработка таблицы vehicle
    existing_cols = {col['name'] for col in inspector.get_columns('vehicle')}
    with db.engine.begin() as conn:
        if 'user_id' not in existing_cols:
            conn.execute(text("ALTER TABLE vehicle ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1"))
            try:
                conn.execute(text("ALTER TABLE vehicle ADD CONSTRAINT fk_vehicle_user_id FOREIGN KEY (user_id) REFERENCES user(id)"))
            except Exception as e:
                print(f"Не удалось добавить внешний ключ для vehicle: {e}")

    # Обработка таблицы driver
    existing_cols = {col['name'] for col in inspector.get_columns('driver')}
    with db.engine.begin() as conn:
        if 'user_id' not in existing_cols:
            conn.execute(text("ALTER TABLE driver ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1"))
            try:
                conn.execute(text("ALTER TABLE driver ADD CONSTRAINT fk_driver_user_id FOREIGN KEY (user_id) REFERENCES user(id)"))
            except Exception as e:
                print(f"Не удалось добавить внешний ключ для driver: {e}")

    # Обработка таблицы cargo_type
    existing_cols = {col['name'] for col in inspector.get_columns('cargo_type')}
    with db.engine.begin() as conn:
        if 'user_id' not in existing_cols:
            conn.execute(text("ALTER TABLE cargo_type ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1"))
            try:
                conn.execute(text("ALTER TABLE cargo_type ADD CONSTRAINT fk_cargo_type_user_id FOREIGN KEY (user_id) REFERENCES user(id)"))
            except Exception as e:
                print(f"Не удалось добавить внешний ключ для cargo_type: {e}")

        # Добавляем колонки для активностей транспорта
        if 'vehicle_activity' not in inspector.get_table_names():
            # Таблица будет создана автоматически через db.create_all()
            pass

with app.app_context():
    db.create_all()
    ensure_route_columns()
    
    # Create default admin user if not exists
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            password=generate_password_hash('admin')
        )
        db.session.add(admin)
        db.session.commit()

if __name__ == '__main__':
    # Создаем админа, если его нет
    with app.app_context():
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                password=generate_password_hash('admin')
            )
            db.session.add(admin)
            db.session.commit()
    
    app.run(debug=True, port=5001) 